"""Expense endpoints: list, create (with ML categorization), delete, CSV/Excel import."""
import csv
import io
from datetime import date, datetime

from openpyxl import load_workbook
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import desc
from sqlalchemy.orm import Session

from ..auth import get_current_user
from ..database import get_db
from ..ml.anomaly import detect_anomalies, detect_duplicates
from ..ml.categorizer import categorize
from ..models import CATEGORIES, Expense, Notification, User
from ..realtime import manager
from ..schemas import CategorizeRequest, CategorizeResponse, ExpenseCreate, ExpenseOut

router = APIRouter(prefix="/api/expenses", tags=["expenses"])


def _refresh_flags(db: Session, user_id: int) -> None:
    """Recompute anomaly/duplicate flags across the user's expenses."""
    expenses = db.query(Expense).filter(Expense.user_id == user_id).all()
    anomalies = detect_anomalies(expenses)
    dups = detect_duplicates(expenses)
    for e in expenses:
        e.is_anomaly = anomalies.get(e.id, False)
        e.is_duplicate = dups.get(e.id, False)
    db.commit()


@router.get("", response_model=list[ExpenseOut])
def list_expenses(
    category: str | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    q = db.query(Expense).filter(Expense.user_id == current.id)
    if category and category != "All":
        q = q.filter(Expense.category == category)
    return q.order_by(desc(Expense.spent_on), desc(Expense.id)).limit(limit).all()


@router.post("", response_model=ExpenseOut, status_code=201)
async def create_expense(
    payload: ExpenseCreate,
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    category = payload.category
    if not category or category == "auto":
        category, _conf = categorize(payload.description)
    elif category not in CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Unknown category '{category}'")

    exp = Expense(
        user_id=current.id,
        description=payload.description,
        amount=payload.amount,
        category=category,
        spent_on=payload.spent_on or date.today(),
    )
    db.add(exp)
    db.commit()
    db.refresh(exp)
    _refresh_flags(db, current.id)
    db.refresh(exp)

    # Real-time overspend alert.
    if payload.amount >= 5000:
        notif = Notification(
            user_id=current.id, kind="overspend",
            title="Large expense recorded",
            body=f"₹{payload.amount:,.0f} on {category} — that's a big one.",
        )
        db.add(notif)
        db.commit()
        await manager.send_to_user(current.id, {
            "type": "notification",
            "kind": "overspend",
            "title": notif.title,
            "body": notif.body,
        })

    await manager.send_to_user(current.id, {"type": "expense_added", "category": category, "amount": payload.amount})
    return exp


@router.delete("/{expense_id}", status_code=204)
def delete_expense(expense_id: int, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    exp = db.query(Expense).filter(Expense.id == expense_id, Expense.user_id == current.id).first()
    if not exp:
        raise HTTPException(status_code=404, detail="Expense not found")
    db.delete(exp)
    db.commit()
    _refresh_flags(db, current.id)


@router.post("/categorize", response_model=CategorizeResponse)
def categorize_endpoint(payload: CategorizeRequest, _: User = Depends(get_current_user)):
    category, confidence = categorize(payload.description)
    return CategorizeResponse(description=payload.description, category=category, confidence=confidence)


def _rows_from_xlsx(raw: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    keys = [str(h or "").strip().lower() for h in header]
    return [
        {k: ("" if v is None else str(v)) for k, v in zip(keys, r)}
        for r in rows
    ]


@router.post("/import", response_model=list[ExpenseOut])
def import_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    """Import expenses from CSV or Excel with columns: date, description, amount[, category]."""
    raw = file.file.read()
    if (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        try:
            raw_rows = _rows_from_xlsx(raw)
        except Exception:
            raise HTTPException(status_code=400, detail="Could not read Excel file")
    else:
        try:
            content = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="File must be UTF-8 encoded CSV or an .xlsx file")
        raw_rows = list(csv.DictReader(io.StringIO(content)))

    created: list[Expense] = []
    for row in raw_rows:
        row = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
        desc = row.get("description") or row.get("desc")
        amt_raw = row.get("amount") or row.get("amt")
        if not desc or not amt_raw:
            continue
        try:
            amount = float(amt_raw.replace(",", "").replace("₹", ""))
        except ValueError:
            continue
        category = row.get("category")
        if not category or category not in CATEGORIES:
            category, _ = categorize(desc)
        spent_on = date.today()
        if row.get("date"):
            row["date"] = row["date"].split(" ")[0]  # strip Excel datetime time part
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    spent_on = datetime.strptime(row["date"], fmt).date()
                    break
                except ValueError:
                    continue
        exp = Expense(user_id=current.id, description=desc, amount=amount, category=category, spent_on=spent_on)
        db.add(exp)
        created.append(exp)

    if not created:
        raise HTTPException(status_code=400, detail="No valid rows found in file")
    db.commit()
    _refresh_flags(db, current.id)
    for e in created:
        db.refresh(e)
    return created
