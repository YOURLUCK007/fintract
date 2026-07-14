"""AI-generated financial reports — JSON preview + PDF/Excel export."""
import io
from collections import defaultdict
from datetime import date

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..auth import get_current_user
from ..database import get_db
from ..ml.health_score import compute_health
from ..ml.recommender import recommendations
from ..ml.subscriptions import summarize as sub_summary
from ..models import Expense, Goal, User

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _build_report(db: Session, current: User) -> dict:
    expenses = db.query(Expense).filter(Expense.user_id == current.id).all()
    goals = db.query(Goal).filter(Goal.user_id == current.id).all()

    months = max(len({f"{e.spent_on.year}-{e.spent_on.month:02d}" for e in expenses}), 1)
    totals: dict[str, float] = defaultdict(float)
    for e in expenses:
        totals[e.category] += e.amount
    total_spend = sum(totals.values())
    monthly_spend = total_spend / months
    monthly_savings = max(current.monthly_income - monthly_spend, 0)

    health = compute_health(current.monthly_income, expenses, goals)
    subs = sub_summary(expenses)
    recs = recommendations(expenses)

    categories = [
        {"category": k, "monthly": round(v / months), "total": round(v)}
        for k, v in sorted(totals.items(), key=lambda kv: kv[1], reverse=True)
    ]

    return {
        "generated_on": date.today().isoformat(),
        "user": current.full_name or current.email,
        "monthly_income": round(current.monthly_income),
        "monthly_spend": round(monthly_spend),
        "monthly_savings": round(monthly_savings),
        "savings_rate": round(monthly_savings / current.monthly_income * 100) if current.monthly_income else 0,
        "health_score": health["score"],
        "categories": categories,
        "subscriptions": subs,
        "recommendations": recs,
    }


@router.get("/summary")
def report_summary(db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    return _build_report(db, current)


def _excel_bytes(rep: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)
    ws["A1"] = "FinTract Financial Report"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Generated on", rep["generated_on"]),
        ("Account", rep["user"]),
        ("Monthly income", rep["monthly_income"]),
        ("Monthly spend", rep["monthly_spend"]),
        ("Monthly savings", rep["monthly_savings"]),
        ("Savings rate (%)", rep["savings_rate"]),
        ("Health score", rep["health_score"]),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws[f"A{i}"] = k
        ws[f"A{i}"].font = bold
        ws[f"B{i}"] = v

    cat = wb.create_sheet("Categories")
    cat.append(["Category", "Monthly (INR)", "Total (INR)"])
    for c in cat[1]:
        c.font = bold
    for row in rep["categories"]:
        cat.append([row["category"], row["monthly"], row["total"]])

    sub = wb.create_sheet("Subscriptions")
    sub.append(["Name", "Cadence", "Amount", "Monthly", "Annual"])
    for c in sub[1]:
        c.font = bold
    for s in rep["subscriptions"]["subscriptions"]:
        sub.append([s["name"], s["cadence"], s["amount"], s["monthly_cost"], s["annual_cost"]])

    rec = wb.create_sheet("Recommendations")
    rec.append(["Recommendation", "Est. monthly saving", "Why"])
    for c in rec[1]:
        c.font = bold
    for r in rep["recommendations"]:
        rec.append([r["title"], r["save"], r["why"]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(width + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_bytes(rep: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title="FinTract Report")
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("FinTract — Financial Report", styles["Title"]))
    story.append(Paragraph(f"Account: {rep['user']} &nbsp;·&nbsp; Generated: {rep['generated_on']}", styles["Normal"]))
    story.append(Spacer(1, 8 * mm))

    summary = [
        ["Monthly income", f"Rs {rep['monthly_income']:,}"],
        ["Monthly spend", f"Rs {rep['monthly_spend']:,}"],
        ["Monthly savings", f"Rs {rep['monthly_savings']:,}"],
        ["Savings rate", f"{rep['savings_rate']}%"],
        ["Financial health score", f"{rep['health_score']}/100"],
    ]
    t = Table(summary, colWidths=[70 * mm, 90 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef1ff")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 6 * mm))

    story.append(Paragraph("Spending by category", styles["Heading2"]))
    cat_data = [["Category", "Monthly (Rs)", "Total (Rs)"]] + [
        [c["category"], f"{c['monthly']:,}", f"{c['total']:,}"] for c in rep["categories"]
    ]
    ct = Table(cat_data, colWidths=[60 * mm, 50 * mm, 50 * mm])
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6c8cff")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(ct)
    story.append(Spacer(1, 6 * mm))

    subs = rep["subscriptions"]["subscriptions"]
    if subs:
        story.append(Paragraph(
            f"Subscriptions detected: {rep['subscriptions']['count']} "
            f"(Rs {rep['subscriptions']['total_monthly']:,}/mo)", styles["Heading2"]))
        sub_data = [["Name", "Cadence", "Monthly (Rs)"]] + [
            [s["name"], s["cadence"], f"{s['monthly_cost']:,}"] for s in subs[:10]
        ]
        st = Table(sub_data, colWidths=[90 * mm, 35 * mm, 35 * mm])
        st.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8a6bff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(st)
        story.append(Spacer(1, 6 * mm))

    story.append(Paragraph("Recommendations", styles["Heading2"]))
    for r in rep["recommendations"]:
        story.append(Paragraph(f"<b>{r['title']}</b> — save ~Rs {r['save']:,}/mo", styles["Normal"]))
        story.append(Paragraph(r["why"], styles["Italic"]))
        story.append(Spacer(1, 2 * mm))

    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(
        "<i>Educational report generated by FinTract. Not guaranteed financial advice.</i>",
        styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


@router.get("/export")
def export_report(
    format: str = Query("pdf", pattern="^(pdf|excel)$"),
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    rep = _build_report(db, current)
    stamp = date.today().isoformat()
    if format == "excel":
        data = _excel_bytes(rep)
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="fintract-report-{stamp}.xlsx"'},
        )
    data = _pdf_bytes(rep)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="fintract-report-{stamp}.pdf"'},
    )
